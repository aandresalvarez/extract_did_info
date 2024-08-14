import yaml
from typing import Dict, Any, Set
from collections import defaultdict
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
import subprocess

# Constants
YAML_FILE_PATH = Path('dby_project.yaml')
EXCEL_OUTPUT_PATH = Path('deid_info_analysis.xlsx')
DATASET_COLUMN_NAMES = {
    "STARR_OMOP_DEID_TEMPLATE": "STARR_OMOP_DEID_TEMPLATE 5.3",
    "AFC_DEID_TEMPLATE_MODERATE_RISK": "AFC_DEID_TEMPLATE_MODERATE_RISK 5.4",
    "AFC_DEID_TEMPLATE_HIGH_RISK": "AFC_DEID_TEMPLATE_HIGH_RISK 5.4 (LDS)",
    "DEID_TEMPLATE": "DEID_TEMPLATE PEDSNET (LDS)",
    "N3C_DEID_TEMPLATE": "N3C_DEID_TEMPLATE 5.3 (LDS)"
}
etl_operations = {
    "add_rand_multiple_offset_specialty": {
        "description": "Implement the root macro AddRandMultipleOperator to add the column deid_col_name's value based on referring column's value<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_1": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_1 in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_2": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_2 in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_3": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_3 in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_4": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_4 in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_5": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_5 in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_adt_event_id": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_ADT_EVENT_ID in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_hsp_account_id": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_HSP_ACCOUNT_ID in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_inpatient_data_id": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_INPATIENT_DATA_ID in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_inpatient_data_id_as_INT64_to_STRING": {
        "description": "Implement the root macro AddRandCastOperator to add the column deid_col_name's value with OFFSET_INPATIENT_DATA_ID in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_note_id": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_NOTE_ID in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_order_id": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_ORDER_ID in codebook table STUDY<br>deid_col_name: string - The target column to DEID"
    },
    "add_rand_offset_pat_enc_csn_id": {
        "description": "Implement the root macro AddRandOperator to add the column deid_col_name's value with OFFSET_PAT_ENC_CSN_ID in codebook table STUDY<br>deid_col_name: string - The target column to DEID cast_as: string - cast as"
    },
    "col_no_deid": {
        "description": "Macro for generating the list of DEID columns that will be used select-statement for except(XXX)<br>afc_deid_template: string - The DEID request template table_id: string - The target table to DEID"
    },
    "del_cnt_DATE": {
        "description": "Implement the root macro DelCntOperator and cast the column as DATE<br>deid_col_name: string - The target column to DEID"
    },
    "del_cnt_DATETIME": {
        "description": "Implement the root macro DelCntOperator and cast the column as DATETIME<br>deid_col_name: string - The target column to DEID"
    },
    "del_cnt_FLOAT64": {
        "description": "Implement the root macro DelCntOperator and cast the column as FLOAT64<br>deid_col_name: string - The target column to DEID"
    },
    "del_cnt_INT64": {
        "description": "Implement the root macro DelCntOperator and cast the column as INT64<br>deid_col_name: string - The target column to DEID"
    },
    "del_cnt_STRING": {
        "description": "Implement the root macro DelCntOperator and cast the column as string<br>deid_col_name: string - The target column to DEID"
    },
    "del_cnt_string": {
        "description": "Implement the root macro DelCntOperator and cast the column as string<br>deid_col_name: string - The target column to DEID"
    },
    "empty_cnt_STRING": {
        "description": "Implement the root macro EmptyCntOperator and cast the column as string<br>deid_col_name: string - The target column to DEID"
    },
    "jitt_date_by_person_id": {
        "description": "Implement the root macro JittDateOperator to jitter deid_col_name with JITTER value in codebook table STUDY_ENTITY (match value by sourcrtable.ID = codebooktable.person_source_value)<br>deid_col_name: string - The target column to DEID"
    },
    "jitt_date_join_by_person_id": {
        "description": "Implement the root macro JittDateJoinOperator to jitter deid_col_name with JITTER value in codebook table STUDY_ENTITY (match value by sourcrtable.ID = codebooktable.person_source_value)<br>deid_col_name: string - The target column to DEID"
    },
    "jitt_date_part_day_by_person_source_value": {
        "description": "Implement the root macro JittDatePartOperator to jitter deid_col_name with JITTER value in codebook table STUDY_ENTITY (match value by sourcrtable.ID = codebooktable.person_source_value)<br>deid_col_name: string - The target column to DEID"
    },
    "jitt_date_part_month_by_person_source_value": {
        "description": "Implement the root macro JittDatePartOperator to jitter deid_col_name with JITTER value in codebook table STUDY_ENTITY (match value by sourcrtable.ID = codebooktable.person_source_value)<br>deid_col_name: string - The target column to DEID"
    },
    "jitt_date_part_year_by_person_source_value": {
        "description": "Implement the root macro JittDatePartOperator to jitter deid_col_name with JITTER value in codebook table STUDY_ENTITY (match value by sourcrtable.ID = codebooktable.person_source_value)<br>deid_col_name: string - The target column to DEID"
    },
    "jitt_datetime_by_person_id": {
        "description": "Implement the root macro JittDateTimeOperator to jitter deid_col_name with JITTER value in codebook table STUDY_ENTITY (match value by sourcrtable.ID = codebooktable.person_source_value)<br>deid_col_name: string - The target column to DEID"
    },
    "jitt_datetime_by_person_source_value": {
        "description": "Implement the root macro JittDateTimeOperator to jitter deid_col_name with JITTER value in codebook table STUDY_ENTITY (match value by sourcrtable.ID = codebooktable.person_source_value)<br>deid_col_name: string - The target column to DEID"
    },
    "jitter_date_day": {
        "description": "Macro for jitter date by a value<br>deid_col_name: string - The target column to be DEID jitter_value: string - The value to jitter"
    },
    "jitter_datetime_day": {
        "description": "Macro for jitter datetime by a value<br>deid_col_name: string - The target column to be DEID jitter_value: string - The value to jitter"
    },
    "reduce_zip_precision": {
        "description": "Macro for reducing zip precision<br>deid_col_name: string - The column to be DEID"
    },
    "remove_day": {
        "description": "Macro for remove day from datetime<br>deid_col_name: string - The target column to be DEID"
    },
    "replace_stanford_reference": {
        "description": "Implement the root macro RegexReplace and replace Stanford reference info<br>deid_col_name: string - The target column to DEID"
    },
    "sub_rand_anon_id_by_person_source_value": {
        "description": "[Not defined]"
    },
    "sub_rand_anon_id_by_provider_source_value": {
        "description": "Implement the root macro SubRandOperator to replace the column deid_col_name with anon_long_id in codebook table STUDY_ENTITY<br>deid_col_name: string - The target column to DEID"
    },
    "sub_rand_anon_long_id_by_person_entity_id": {
        "description": "Implement the root macro SubRandOperator to replace the column deid_col_name with anon_long_id in codebook table STUDY_ENTITY<br>deid_col_name: string - The target column to DEID"
    },
    "sub_rand_anon_long_id_by_person_id": {
        "description": "Implement the root macro SubRandOperator to replace the column deid_col_name with anon_long_id in codebook table STUDY_ENTITY<br>deid_col_name: string - The target column to DEID"
    },
}


# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DeIdConfigError(Exception):
    """Custom exception for de-identification configuration errors."""
    pass

class YamlHandler:
    @staticmethod
    def fix_indentation(yaml_file: Path) -> None:
        try:
            subprocess.run(['expand', '-t', '2', str(yaml_file)], check=True, capture_output=True, text=True)
            logger.info(f"YAML indentation fixed for: {yaml_file}")
        except FileNotFoundError:
            logger.error("The 'expand' command was not found. Please ensure it is installed.")
        except subprocess.CalledProcessError as e:
            logger.error(f"Error fixing YAML indentation: {e}")

    @staticmethod
    def load_yaml(yaml_file: Path) -> Dict[str, Any]:
        YamlHandler.fix_indentation(yaml_file)
        try:
            with open(yaml_file, 'r') as file:
                return yaml.safe_load(file)
        except FileNotFoundError:
            logger.error(f"YAML file not found: {yaml_file}")
            raise
        except yaml.YAMLError as e:
            logger.error(f"Error parsing YAML file: {e}")
            raise

class DeIdInfoExtractor:
    @staticmethod
    def extract(config: Dict[str, Any]) -> Dict[str, Dict[str, Dict[str, str]]]:
        deid_info = defaultdict(lambda: defaultdict(dict))

        def process_section(section_name: str, section_data: Any) -> None:
            if isinstance(section_data, dict):
                if 'tables_to_deid' in section_data:
                    dataset_name = section_name.upper()
                    for table_config in section_data['tables_to_deid']:
                        DeIdInfoExtractor._process_table(deid_info, dataset_name, table_config)
                else:
                    for nested_section_name, nested_section_data in section_data.items():
                        process_section(nested_section_name, nested_section_data)

        for section_name, section_data in config.items():
            process_section(section_name, section_data)

        return dict(deid_info)

    @staticmethod
    def _process_table(deid_info: Dict[str, Dict[str, Dict[str, str]]], dataset_name: str, table_config: Dict[str, Any]) -> None:
        try:
            table_id = table_config['table_id']
            table_info = deid_info[table_id][dataset_name]

            for operation in table_config.get('col_deid_operations', []):
                table_info[operation['col_id'].lower()] = operation['op_name']

            for no_op_col in table_config.get('col_no_deid', []):
                table_info[no_op_col['col_id'].lower()] = 'col_no_deid'
        except KeyError as e:
            raise DeIdConfigError(f"Missing required key in table configuration: {e}")

class DeIdAnalyzer:
    @staticmethod
    def get_distinct_values(deid_info: Dict[str, Dict[str, Dict[str, str]]]) -> Set[str]:
        return set(value for table_data in deid_info.values() for dataset_data in table_data.values() for value in dataset_data.values())

class ExcelExporter:
    def __init__(self, deid_info: Dict[str, Dict[str, Dict[str, str]]], output_file: Path):
        self.deid_info = deid_info
        self.output_file = output_file
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.styles = self._define_styles()

    def export(self):
        distinct_values = DeIdAnalyzer.get_distinct_values(self.deid_info)
        self._create_distinct_values_sheet(distinct_values)
        self._create_data_sheets()
        self.wb.save(self.output_file)
        logger.info(f"Read-only Excel file created with distinct values sheet and improved formatting: {self.output_file}")

    def _define_styles(self):
        return {
            'header': {
                'font': Font(name='Calibri', size=12, bold=True, color="FFFFFF"),
                'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            },
            'body': {
                'font': Font(name='Calibri', size=11),
                'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
                'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            }
        }

    def _create_distinct_values_sheet(self, distinct_values: Set[str]):
        ws = self.wb.create_sheet(title="DEID Operations", index=0)

        # Set header
        ws.append(["Operation", "Description"])
        for cell in ws[1]:
            self._apply_style(cell, self.styles['header'])

        # Populate data from etl_operations dictionary
        for operation_name, operation_data in etl_operations.items():
            ws.append([operation_name, operation_data['description']])

        # Apply body style to all data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                self._apply_style(cell, self.styles['body'])

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Freeze panes and protect sheet
        ws.freeze_panes = 'A2'
        ws.protection = SheetProtection(sheet=True, formatColumns=False, formatRows=False, sort=False, autoFilter=False)

        # Allow editing only in the "Description" column
        for row in ws.iter_rows(min_row=2):
            row[1].protection = Protection(locked=False)  # Unlock the second cell (Description)
    
    def _create_data_sheets(self):
        # Sort table names alphabetically
        sorted_table_names = sorted(self.deid_info.keys())
        
        for table_name in sorted_table_names:
            table_data = self.deid_info[table_name]
            ws = self.wb.create_sheet(title=table_name)
            datasets = list(DATASET_COLUMN_NAMES.keys())
            header = ['Column'] + [DATASET_COLUMN_NAMES[ds] for ds in datasets]
            ws.append(header)
            self._apply_row_style(ws[1], self.styles['header'])

            column_mapping = {column.lower(): column for dataset in datasets for column in table_data[dataset]}

            for column in sorted(column_mapping.values()):
                row = [column] + [table_data[dataset].get(column, 'N/A') for dataset in datasets]
                ws.append(row)

            for row in ws.iter_rows(min_row=2):
                self._apply_row_style(row, self.styles['body'])

            self._adjust_column_widths(ws)
            ws.freeze_panes = 'A2'
            ws.protection = SheetProtection(sheet=True, formatColumns=False, formatRows=False, sort=False, autoFilter=False)

    @staticmethod
    def _apply_style(cell, style):
        for attr, value in style.items():
            setattr(cell, attr, value)

    def _apply_row_style(self, row, style):
        for cell in row:
            self._apply_style(cell, style)

    @staticmethod
    def _adjust_column_widths(ws):
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(adjusted_width, 50)

def main():
    try:
        config = YamlHandler.load_yaml(YAML_FILE_PATH)
        deid_data = DeIdInfoExtractor.extract(config)
        exporter = ExcelExporter(deid_data, EXCEL_OUTPUT_PATH)
        exporter.export()
        print(f"De-identification data exported to: {EXCEL_OUTPUT_PATH}")
    except (FileNotFoundError, yaml.YAMLError, DeIdConfigError) as e:
        logger.error(f"An error occurred: {e}")

if __name__ == "__main__":
    main()